import openpyxl
import numpy as np

PATH = "MainExcelFile.xlsx"
N = 300
NUMBER_OF_SHIFTS = 10 # значение должно быть меньше чем N хотя бы на 3 

def get_values_from_xlsx():
    specified_set = []
    generated_set = []
    wb_obj = openpyxl.load_workbook(PATH)
    for i in range(1, N + 1):
        specified_set.append(wb_obj.active.cell(row = i, column = 2).value)
        generated_set.append(wb_obj.active.cell(row = i, column = 3).value)
    return specified_set, generated_set


def find_corell_coefs(values):
    ret_arr = []
    for i in range(1, min(NUMBER_OF_SHIFTS + 1, N - 2)): 
        selection1 = values[:-i]
        selection2 = values[i:]
        mean1 = np.mean(selection1)
        mean2 = np.mean(selection2)
        top = 0
        bottom_left = 0
        bottom_right = 0
        for j in range(len(selection1)):
            curr_diff1 = selection1[j] - mean1
            curr_deff2 = selection2[j] - mean2
            top += curr_diff1 * curr_deff2
            bottom_left += curr_diff1**2
            bottom_right += curr_deff2**2
        coef = top / np.sqrt(bottom_left * bottom_right)
        ret_arr.append(coef.item())
    return ret_arr

def print_values(values, header):
    print("----- " + header + " ------")
    for i in range(len(values)):
        print(i + 1, ":", round(values[i], 4))
    print("Max:", max(values))
    print("Min:", min(values))
    print("----------------------------\n")

def solve(specified_set, generated_set):
    spec = find_corell_coefs(specified_set)
    gen = find_corell_coefs(generated_set)
    print_values(spec, "Заданная числовая последовательность")
    print_values(gen, "Сгенерированная числовая последовательность")
    print_values(abs((np.array(spec) - np.array(gen)) / np.array(gen))  * 100, "Процентное отношение")